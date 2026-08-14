"""
Document Reader and Management Engine for ITC Brand Marketing.
Directly reads and writes marketing documents (Campaign Hooks, Creative Hooks, Media Plans, Audience Segments, Brand Guidelines)
without requiring vector embeddings.
"""

import os
import sys
from typing import Dict, List, Any, Optional
import pypdf

# Define Base Paths with Multi-Environment Fallbacks
AGENT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ROOT_DIR = os.path.dirname(AGENT_DIR)

def _resolve_dir(subpath: str) -> str:
    candidates = [
        os.path.join(ROOT_DIR, subpath),
        os.path.join(AGENT_DIR, subpath),
        os.path.join("/app", subpath),
        os.path.join(os.getcwd(), subpath)
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    # Fallback to creating under ROOT_DIR or /tmp
    primary = os.path.join(ROOT_DIR, subpath)
    try:
        os.makedirs(primary, exist_ok=True)
        return primary
    except Exception:
        tmp_dir = os.path.join("/tmp", subpath)
        os.makedirs(tmp_dir, exist_ok=True)
        return tmp_dir

ITC_MARKETING_DIR = _resolve_dir(os.path.join("ITC Marketing", "ITC Marketing Files"))
IAB_FORMATS_DIR = _resolve_dir("IAB Formats")
GENERATED_ASSETS_DIR = _resolve_dir("generated_assets")

os.makedirs(os.path.join(GENERATED_ASSETS_DIR, "images"), exist_ok=True)
os.makedirs(os.path.join(GENERATED_ASSETS_DIR, "videos"), exist_ok=True)
os.makedirs(os.path.join(GENERATED_ASSETS_DIR, "reports"), exist_ok=True)


def list_marketing_folders() -> Dict[str, Any]:
    """Lists all available folders and files in the ITC Marketing and IAB Formats directories."""
    result = {"itc_marketing_files": {}, "iab_formats": []}
    try:
        if os.path.exists(ITC_MARKETING_DIR):
            for root, dirs, files in os.walk(ITC_MARKETING_DIR):
                folder_name = os.path.relpath(root, ITC_MARKETING_DIR)
                result["itc_marketing_files"][folder_name] = [f for f in files if not f.startswith('.')]
        if os.path.exists(IAB_FORMATS_DIR):
            result["iab_formats"] = [f for f in os.listdir(IAB_FORMATS_DIR) if not f.startswith('.')]
        return result
    except Exception as e:
        return {"error": str(e)}


def read_marketing_document(folder_name: str, file_name: str) -> str:
    """
    Reads a marketing document (Campaign brief, Creative hooks, Media plan, Brand guidelines, Analytics CSV, or Audience Excel).
    
    Args:
        folder_name: Subfolder name (e.g. 'Campaign Hooks', 'Creative Hooks', 'Media Plan', 'Brand Guidelines', 'Audience', 'Historical campaign and channel performance').
        file_name: Exact file name (e.g. 'fiama_campaign_brief.pdf', 'itc_limited_brand_guidelines_2026.md', 'itc_customer_segments_demo.xlsx').
    """
    # Check both direct match and normalized match
    file_path = os.path.join(ITC_MARKETING_DIR, folder_name, file_name)
    if not os.path.exists(file_path):
        # Search recursively for the filename
        for root, dirs, files in os.walk(ITC_MARKETING_DIR):
            for f in files:
                if f.lower() == file_name.lower():
                    file_path = os.path.join(root, f)
                    break

    if not os.path.exists(file_path):
        return f"Error: Document '{file_name}' not found in '{folder_name}'."

    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".pdf":
            text = []
            with open(file_path, "rb") as f:
                reader = pypdf.PdfReader(f)
                for i, page in enumerate(reader.pages):
                    extracted = page.extract_text()
                    if extracted:
                        text.append(f"--- Page {i+1} ---\n" + extracted)
            return "\n\n".join(text) if text else "PDF content is empty or unreadable."
        elif ext in [".xlsx", ".xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            output = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                output.append(f"### Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    if any(row):
                        output.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
            return "\n".join(output)
        elif ext == ".csv":
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        else:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
    except Exception as e:
        return f"Error reading document '{file_name}': {e}"


def save_marketing_document(folder_name: str, file_name: str, content: str) -> str:
    """
    Saves a generated marketing document into the ITC Marketing Files directory.
    
    Args:
        folder_name: Target subfolder (e.g. 'Campaign Hooks', 'Creative Hooks', 'Media Plan').
        file_name: Target filename (e.g. 'sunfeast_dark_fantasy_campaign_brief.md').
        content: The text/markdown content to write.
    """
    target_dir = os.path.join(ITC_MARKETING_DIR, folder_name)
    os.makedirs(target_dir, exist_ok=True)
    target_path = os.path.join(target_dir, file_name)
    try:
        with open(target_path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"Document successfully saved to: {target_path}"
    except Exception as e:
        return f"Error saving document: {e}"


def read_iab_guidelines() -> str:
    """Reads available IAB format specs and guidelines."""
    try:
        pdf_path = os.path.join(IAB_FORMATS_DIR, "IABNewAdPortfolio_LW_FixedSizeSpec.pdf")
        if os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                reader = pypdf.PdfReader(f)
                return "\n".join([p.extract_text() for p in reader.pages if p.extract_text()])
        files = os.listdir(IAB_FORMATS_DIR)
        return f"IAB Format files available: {files}. Standard IAB sizing and LEAN rules apply."
    except Exception as e:
        return f"Error reading IAB guidelines: {e}"
